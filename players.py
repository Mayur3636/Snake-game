import openpyxl


class Players:

    def __init__(self):
        self.all_data = openpyxl.load_workbook("players.xlsx")
        self.player_data = self.all_data["players"]
        self.current_player = 0

    def is_player_in_data(self, player):
        row = self.player_data.max_row
        for i in range(1, row + 1):
            if self.player_data.cell(row=i, column=1).value == player:
                return True
        return False

    def save_data(self):
        self.all_data.save("players.xlsx")

    def add_player(self, player):
        r = self.player_data.max_row + 1
        self.player_data.cell(row=r, column=1, value=player)
        self.player_data.cell(row=r, column=2, value=0)
        self.player_data.cell(row=r, column=3, value=0)
        self.save_data()

    def get_player(self, player):
        row = self.player_data.max_row
        m = []
        for i in range(1, row + 1):
            if self.player_data.cell(row=i, column=1).value == player:
                m.append(self.player_data.cell(row=i, column=1).value)
                m.append(self.player_data.cell(row=i, column=2).value)
                m.append(self.player_data.cell(row=i, column=3).value)
        return m

    def update_player(self, player_list):
        name = player_list[0]
        score = player_list[1]
        row = self.player_data.max_row
        for i in range(1, row + 1):
            if self.player_data.cell(row=i, column=1).value == name:
                self.player_data.cell(row=i, column=2, value=score)
        self.save_data()

    def high_score(self, player_list, curr_score):
        row = self.player_data.max_row
        player = player_list[0]
        for i in range(1, row + 1):
            if self.player_data.cell(row=i, column=1).value == player:
                # self.player_data.cell(row=i, column=2, value=curr_score)
                if curr_score >= self.player_data.cell(row=i, column=3).value:
                    self.player_data.cell(row=i, column=3, value=curr_score)
        self.save_data()

    def get_high_score(self, player_list):
        row = self.player_data.max_row
        player = player_list[0]
        for i in range(1, row + 1):
            if self.player_data.cell(row=i, column=1).value == player:
                m = self.player_data.cell(row=i, column=3).value
        return m